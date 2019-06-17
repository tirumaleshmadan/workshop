import pymysql
pymysql.install_as_MySQLdb()
import MySQLdb
import click
import openpyxl
import os.path
import dateutil.parser

@click.group()
def connect():
      pass

@connect.command()
@click.argument("dbname",nargs=1)
@click.argument("tablename",nargs=1)
@click.argument("excelfile",type=click.Path(exists=True))
def importdata(dbname,tablename,excelfile):
      excelfile=os.path.basename(excelfile)
      db = MySQLdb.connect(host='127.0.0.1', port=3306, user="root", passwd="Meghana123###",db=dbname)
      c = db.cursor()
      c.execute(f"use {dbname}")
      workbook = openpyxl.load_workbook(excelfile)
      sheet = workbook.active
      max_col = sheet.max_column
      max_row = sheet.max_row
      for i in range(2, max_row+1 ):
            print(i)
            li=[]
            li.append(i-1)
            for j in range(1, max_col + 1):
                  cells = sheet.cell(row=i, column=j)
                  if cells.value==None:
                      li.append("")
                  else:
                      li.append(cells.value)
            li=tuple(li)
            print(li)
            Query=f"insert into {tablename} values {li};"
            c.execute(Query)
      db.commit()

if __name__ == '__main__':
      connect()